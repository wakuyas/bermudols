#!/usr/bin/env python3
"""
generate_season.py
──────────────────
Bmds_shukei_XXXX.xlsx を読み込んで season_YYYY.html を生成します。

使い方:
    python generate_season.py ../Bermudas_works/Bmds_shukei_2026.xlsx 2026
    python generate_season.py ../Bermudas_works/Bmds_shukei_2025.xlsx 2025

生成ファイル: season_YYYY.html（例: season_2026.html）
season.html は常に最新年度へのリダイレクトページです。
"""

import sys, json, re
from pathlib import Path
import openpyxl

# ─── helpers ──────────────────────────────────────────────────────────────

def cell_str(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()

def ip_to_num(ip_str):
    """投球回文字列 → 数値  例: '5.1'→5.333, '5.2'→5.667, '5 1/3'→5.333, '5'→5.0"""
    s = str(ip_str).strip()
    if not s:
        return 0.0
    if "1/3" in s:
        return int(s.split()[0]) + 1/3
    if "2/3" in s:
        return int(s.split()[0]) + 2/3
    if "." in s:
        w, f = s.split(".", 1)
        return int(w) + int(f) / 3
    return float(s)

def num_to_ip_str(n):
    """数値 → 投球回表示文字列  例: 5.333 → '5 1/3'"""
    w = int(n)
    frac = n - w
    if abs(frac - 1/3) < 0.01:
        return f"{w} 1/3"
    if abs(frac - 2/3) < 0.01:
        return f"{w} 2/3"
    return str(w)

def fmt_rate(val):
    if val != val or val is None:  # nan check
        return "—"
    s = f"{val:.3f}"
    return s[1:] if s.startswith("0.") else s  # strip leading zero

# ─── Excel 読み込み ────────────────────────────────────────────────────────

def load_excel(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── 試合一覧 ──────────────────────────────────────────────────────────
    ws_games = wb["試合一覧"]
    games = []  # list of dicts
    for row in ws_games.iter_rows(min_row=3, values_only=True):
        date, opp, result, bd, opp_score = (row[i] if i < len(row) else None for i in range(5))
        if not date or not opp:
            continue
        innings = []
        for i in range(5, 14):
            v = row[i] if i < len(row) else None
            innings.append("" if v is None else str(v))
        games.append({
            "date": str(date).strip(),
            "opp":  str(opp).strip(),
            "result": str(result).strip() if result else "",
            "bd":   int(bd)   if bd   is not None else 0,
            "opp_score": int(opp_score) if opp_score is not None else 0,
            "innings": innings,
        })

    # ── 打撃成績 ──────────────────────────────────────────────────────────
    ws_bat = wb["打撃成績"]
    game_bat   = {}   # {opp: [player_row, ...]}
    player_bat = {}   # {name: [game_row, ...]}
    for row in ws_bat.iter_rows(min_row=3, values_only=True):
        if len(row) < 17:
            continue
        date, opp, name = (str(row[i]).strip() if row[i] else "" for i in range(3))
        if not opp or not name:
            continue
        def v(i):
            x = row[i] if i < len(row) else None
            return "" if x is None else str(int(x)) if isinstance(x, float) and x == int(x) else str(x).strip()
        pa,ab,r,h1,h2,h3,hr,rbi,sb,sac,sf,bb,so,lob = (v(i) for i in range(3,17))
        h = str((int(h1 or 0)+int(h2 or 0)+int(h3 or 0)+int(hr or 0)))
        rec = {"date":date,"opp":opp,"name":name,
               "pa":pa,"ab":ab,"r":r,"h1":h1,"h2":h2,"h3":h3,"hr":hr,
               "h":h,"rbi":rbi,"sb":sb,"sac":sac,"sf":sf,"bb":bb,"so":so,"lob":lob}
        game_bat.setdefault(f"{opp}|{date}", []).append(rec)
        player_bat.setdefault(name, []).append({**rec})

    # ── 投手成績 ──────────────────────────────────────────────────────────
    ws_pit = wb["投手成績"]
    game_pit   = {}   # {"opp|date": [...]}
    player_pit = {}
    for row in ws_pit.iter_rows(min_row=3, values_only=True):
        if len(row) < 11:
            continue
        date, opp, name, wl = (str(row[i]).strip() if row[i] else "" for i in range(4))
        if not opp or not name:
            continue
        def pv(i):
            x = row[i] if i < len(row) else None
            return "" if x is None else str(x).strip()
        ip,bf,h,bb,so,r,er = (pv(i) for i in range(4,11))
        # normalize wl
        wl_norm = {"○":"○","●":"●","S":"S","勝":"○","敗":"●","s":"S"}.get(wl, wl)
        rec = {"date":date,"opp":opp,"name":name,"wl":wl_norm,
               "ip":ip,"bf":bf,"h":h,"bb":bb,"so":so,"r":r,"er":er}
        game_pit.setdefault(f"{opp}|{date}", []).append(rec)
        player_pit.setdefault(name, []).append({**rec})

    return games, game_bat, player_bat, game_pit, player_pit

# ─── 集計 ─────────────────────────────────────────────────────────────────

def compute_team_stats(games, game_bat, game_pit):
    wins = sum(1 for g in games if g["result"] == "勝")
    losses = sum(1 for g in games if g["result"] == "負")
    draws = sum(1 for g in games if g["result"] == "分")
    total = len(games)

    # チーム打撃集計
    tb = {"r":0,"h":0,"rbi":0,"sb":0,"bb":0,"so":0,"ab":0,"pa":0,"h2":0,"h3":0,"hr":0}
    for recs in game_bat.values():
        for r in recs:
            for k in tb:
                tb[k] += int(r.get(k) or 0)
    avg  = fmt_rate(tb["h"]/tb["ab"])  if tb["ab"] else "—"
    obp  = fmt_rate((tb["h"]+tb["bb"])/(tb["ab"]+tb["bb"])) if (tb["ab"]+tb["bb"]) else "—"
    slg_num = tb["h"] + tb["h2"] + 2*tb["h3"] + 3*tb["hr"]
    slg  = fmt_rate(slg_num/tb["ab"]) if tb["ab"] else "—"

    # チーム投手集計
    tp = {"w":0,"l":0,"s":0,"ip":0.0,"h":0,"bb":0,"so":0,"r":0,"er":0}
    for recs in game_pit.values():
        for r in recs:
            if r.get("wl") == "○": tp["w"] += 1
            if r.get("wl") == "●": tp["l"] += 1
            if r.get("wl") == "S": tp["s"] += 1
            tp["ip"] += ip_to_num(r.get("ip","0"))
            for k in ["h","bb","so","r","er"]:
                tp[k] += int(r.get(k) or 0)
    era = fmt_rate(tp["er"]*9/tp["ip"]) if tp["ip"] else "—"

    return wins, losses, draws, total, tb, avg, obp, slg, tp, era

def compute_player_bat_stats(player_bat, total_games=0):
    qual_pa = total_games * 1.3
    stats = []
    for name, games in player_bat.items():
        g = len(games)
        ab=h=r=rbi=sb=bb=so=pa=h2=h3=hr=sac=sf=lob = 0
        for rec in games:
            pa  += int(rec.get("pa") or 0)
            ab  += int(rec.get("ab") or 0)
            r   += int(rec.get("r")  or 0)
            h   += int(rec.get("h")  or 0)
            h2  += int(rec.get("h2") or 0)
            h3  += int(rec.get("h3") or 0)
            hr  += int(rec.get("hr") or 0)
            rbi += int(rec.get("rbi")or 0)
            sb  += int(rec.get("sb") or 0)
            bb  += int(rec.get("bb") or 0)
            so  += int(rec.get("so") or 0)
            sac += int(rec.get("sac")or 0)
            sf  += int(rec.get("sf") or 0)
            lob += int(rec.get("lob")or 0)
        avg = fmt_rate(h/ab)            if ab else "—"
        obp = fmt_rate((h+bb)/(ab+bb)) if (ab+bb) else "—"
        slg = fmt_rate((h+h2+2*h3+3*hr)/ab) if ab else "—"
        ops_val = None
        try:
            ops_val = float(obp.replace("—","")) + float(slg.replace("—",""))
        except:
            pass
        ops = fmt_rate(ops_val) if ops_val is not None else "—"
        try:
            avg_num = float(avg) if avg != "—" else -1.0
        except:
            avg_num = -1.0
        qualified = (pa >= qual_pa) if qual_pa > 0 else False
        stats.append({"name":name,"g":g,"pa":pa,"ab":ab,"r":r,"h":h,
                       "h2":h2,"h3":h3,"hr":hr,"rbi":rbi,"sb":sb,
                       "bb":bb,"so":so,"avg":avg,"avg_num":avg_num,
                       "obp":obp,"slg":slg,"ops":ops,"qual":qualified})
    # 規定到達グループを先に、各グループ内は打率降順
    stats.sort(key=lambda x: (0 if x["qual"] else 1, -x["avg_num"]))
    return stats

def compute_player_pit_stats(player_pit, total_games):
    qual_ip = total_games * 1.0  # 資格投球回: 試合数×1
    stats = []
    for name, games in player_pit.items():
        g = len(games)
        w=l=s=bf=h=bb=so=r=er = 0
        ip = 0.0
        for rec in games:
            if rec.get("wl") == "○": w += 1
            if rec.get("wl") == "●": l += 1
            if rec.get("wl") == "S": s += 1
            ip += ip_to_num(rec.get("ip","0"))
            bf += int(rec.get("bf") or 0)
            h  += int(rec.get("h")  or 0)
            bb += int(rec.get("bb") or 0)
            so += int(rec.get("so") or 0)
            r  += int(rec.get("r")  or 0)
            er += int(rec.get("er") or 0)
        era_val = er*9/ip if ip else None
        era = fmt_rate(era_val) if era_val is not None else "—"
        ip_disp = num_to_ip_str(ip)
        qualified = ip >= qual_ip
        stats.append({"name":name,"g":g,"w":w,"l":l,"s":s,
                       "ip":ip_disp,"ip_num":ip,"bf":bf,"h":h,
                       "bb":bb,"so":so,"r":r,"er":er,"era":era,
                       "era_num": era_val if era_val is not None else float("inf"),
                       "qual":qualified})
    # 規定到達グループを先に、各グループ内は防御率昇順（良い順）
    stats.sort(key=lambda x: (0 if x["qual"] else 1, x["era_num"]))
    return stats

# ─── HTML 生成 ─────────────────────────────────────────────────────────────

RESULT_CLASS = {"勝":"win","負":"loss","分":"draw"}
RESULT_LABEL = {"勝":"勝","負":"負","分":"分"}

def game_rows_html(games):
    rows = []
    for g in games:
        rc  = RESULT_CLASS.get(g["result"], "")
        lbl = RESULT_LABEL.get(g["result"], g["result"])
        inn = "".join(
            f'<td class="inn-cell">{v if v else "·"}</td>' for v in g["innings"]
        )
        opp_esc  = g["opp"].replace("'", "\\'")
        date_esc = g["date"].replace("'", "\\'")
        rows.append(
            f'<tr class="game-row {rc}">'
            f'<td class="date-col"><a class="game-link" href="#" onclick="showGame(event,\'{opp_esc}\',\'{date_esc}\')">{g["date"]}</a></td>'
            f'<td class="opp-col"><a class="game-link" href="#" onclick="showGame(event,\'{opp_esc}\',\'{date_esc}\')">{g["opp"]}</a></td>'
            f'<td class="score-cell bd-score">{g["bd"]}</td>'
            f'<td class="score-sep">-</td>'
            f'<td class="score-cell opp-score">{g["opp_score"]}</td>'
            f'<td class="result-badge-cell"><span class="result-badge {rc}">{lbl}</span></td>'
            f'{inn}</tr>'
        )
    return "\n".join(rows)

def bat_rows_html(stats, total_games):
    rows = []
    qual_section = False
    unqual_section = False
    for p in stats:
        q = p.get("qual", False)
        if q and not qual_section:
            rows.append(f'<tr class="section-divider"><td colspan="17">◎ 規定打席到達（{total_games} 試合 × 1.3 打席以上）</td></tr>')
            qual_section = True
        if not q and not unqual_section:
            rows.append('<tr class="section-divider"><td colspan="17">○ 規定打席未満</td></tr>')
            unqual_section = True
        cls = "" if q else " class=\"unqual\""
        ne = p["name"].replace("'", "\\'")
        rows.append(
            f'<tr{cls}>'
            f'<td class="name-col"><a class="player-link" href="#" onclick="showPlayer(event,\'{ne}\')">{p["name"]}</a></td>'
            f'<td>{p["g"]}</td><td>{p["pa"]}</td><td>{p["ab"]}</td>'
            f'<td>{p["r"]}</td><td class="hit-col">{p["h"]}</td>'
            f'<td>{p["rbi"]}</td><td>{p["sb"]}</td>'
            f'<td>{p["h2"]}</td><td>{p["h3"]}</td><td>{p["hr"]}</td>'
            f'<td>{p["bb"]}</td><td>{p["so"]}</td>'
            f'<td class="rate-col">{p["avg"]}</td>'
            f'<td class="rate-col">{p["obp"]}</td>'
            f'<td class="rate-col">{p["slg"]}</td>'
            f'<td class="rate-col ops-col">{p["ops"]}</td>'
            f'</tr>'
        )
    return "\n".join(rows)

def pit_rows_html(stats, total_games):
    rows = []
    for p in stats:
        q = p["qual"]
        badge = '<span class="qual-badge">規定</span>' if q else ""
        ne = p["name"].replace("'", "\\'")
        rows.append(
            f'<tr>'
            f'<td class="name-col"><a class="player-link" href="#" onclick="showPlayer(event,\'{ne}\')">{p["name"]}</a>{badge}</td>'
            f'<td>{p["g"]}</td><td>{p["w"]}</td><td>{p["l"]}</td><td>{p["s"]}</td>'
            f'<td class="ip-col">{p["ip"]}</td><td>{p["bf"]}</td><td>{p["h"]}</td>'
            f'<td>{p["bb"]}</td><td>{p["so"]}</td><td>{p["r"]}</td><td>{p["er"]}</td>'
            f'<td class="rate-col">{p["era"]}</td>'
            f'</tr>'
        )
    return "\n".join(rows)

def ranking_sections_html(bat_stats, pit_stats):
    def rank_items(players, key, label, fmt=str, reverse=True):
        valid = [(p["name"], p[key]) for p in players if isinstance(p[key], (int,float)) or (isinstance(p[key], str) and p[key] not in ("—",""))]
        try:
            valid2 = [(n, float(v)) for n,v in valid]
        except:
            valid2 = []
        valid2.sort(key=lambda x: -x[1] if reverse else x[1])
        html = f'<div class="rank-card"><div class="rank-hdr">{label}</div>'
        for i, (nm, val) in enumerate(valid2[:5], 1):
            gold = ' rank-1' if i == 1 else ''
            html += (
                f'<div class="rank-row{gold}">'
                f'<span class="rank-num">{i}</span>'
                f'<span class="rank-name">{nm}</span>'
                f'<span class="rank-val">{fmt(val)}</span>'
                f'</div>'
            )
        html += '</div>'
        return html

    def fmt_avg(v):
        s = f"{float(v):.3f}"
        return s[1:] if s.startswith("0.") else s

    sections = '<div class="rank-grid">'
    sections += rank_items(bat_stats, "h",   "安打",  int)
    sections += rank_items(bat_stats, "hr",  "本塁打", int)
    sections += rank_items(bat_stats, "rbi", "打点",  int)
    sections += rank_items(bat_stats, "sb",  "盗塁",  int)
    sections += rank_items(pit_stats, "w",   "投手勝利", int)
    sections += rank_items(pit_stats, "so",  "奪三振", int)
    sections += '</div>'
    return sections

# ─── HTML テンプレート ─────────────────────────────────────────────────────

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>バミューダドルフィンズ {year}年 シーズン成績</title>
<style>
  :root {{
    --primary: #1a3a5c; --accent: #c8102e; --gold: #d4a017;
    --bg: #f5f0e8; --card-bg: #fffef9; --border: #d4c8a8;
    --text: #2c2416; --muted: #7a6e5a;
    --win-bg: #e6f4ec; --win-cl: #1a6e3a;
    --loss-bg: #fce8e8; --loss-cl: #c8102e;
    --draw-bg: #fef9e7; --draw-cl: #b07d10;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Hiragino Sans','Hiragino Kaku Gothic ProN','Noto Sans JP',sans-serif;
         background: var(--bg); color: var(--text); min-height: 100vh; }}
  .nav {{ background: var(--primary); padding: 10px 24px; display: flex; gap: 6px;
          flex-wrap: wrap; align-items: center; }}
  .nav a {{ color: #a0b8d0; text-decoration: none; font-size: 13px; padding: 4px 10px;
            border-radius: 4px; transition: all .15s; }}
  .nav a:hover, .nav a.active {{ background: rgba(255,255,255,.12); color: #fff; }}
  .hero {{ background: linear-gradient(160deg,#0d2238 0%,#1a3a5c 55%,#1e4d7a 100%);
           color: #fff; padding: 36px 24px 28px; text-align: center; position: relative; overflow: hidden; }}
  .hero::before {{ content: '⚾'; position: absolute; font-size: 280px; opacity: .04;
                   top: -40px; right: -20px; line-height: 1; pointer-events: none; }}
  .hero-year  {{ font-size: 12px; letter-spacing: .25em; color: #a0b8d0; margin-bottom: 8px; }}
  .hero-title {{ font-size: clamp(24px,5vw,42px); font-weight: 900; margin-bottom: 20px; }}
  .hero-title span {{ color: var(--gold); }}
  .rec-row {{ display: flex; justify-content: center; gap: 0; }}
  .rec-item {{ text-align: center; padding: 10px 22px;
               border-right: 1px solid rgba(255,255,255,.15); }}
  .rec-item:last-child {{ border-right: none; }}
  .rec-num {{ font-size: 40px; font-weight: 900; line-height: 1; }}
  .rec-lbl {{ font-size: 12px; color: #a0b8d0; margin-top: 4px; }}
  .rec-win  {{ color: #5fcc88; }} .rec-loss {{ color: #ff7a85; }} .rec-draw {{ color: var(--gold); }}
  .main {{ max-width: 1100px; margin: 0 auto; padding: 32px 24px 60px; }}
  .sec-hdr {{ font-size: 18px; font-weight: 900; color: var(--primary);
              border-left: 4px solid var(--primary); padding-left: 12px;
              margin: 32px 0 14px; }}
  .team-card {{ background: var(--card-bg); border: 1px solid var(--border);
                border-radius: 12px; overflow: hidden; margin-bottom: 28px; }}
  .tc-head {{ background: var(--primary); color: #fff; padding: 10px 16px;
              font-weight: 700; font-size: 14px; }}
  .tc-stats {{ display: flex; flex-wrap: wrap; }}
  .tc-stat {{ padding: 14px 20px; text-align: center;
              border-right: 1px solid var(--border); border-bottom: 1px solid var(--border); }}
  .tc-val  {{ font-size: 22px; font-weight: 900; color: var(--primary); }}
  .tc-lbl  {{ font-size: 10px; color: var(--muted); margin-top: 2px; }}
  .tc-rate {{ font-size: 16px; font-weight: 900; color: var(--primary); }}
  .games-wrap {{ overflow-x: auto; }}
  table.games-tbl {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  table.games-tbl th {{ background: var(--primary); color: #fff; padding: 8px 10px;
                        text-align: center; white-space: nowrap; }}
  table.games-tbl td {{ padding: 7px 10px; border-bottom: 1px solid #ede8dc; text-align: center; }}
  table.games-tbl tr:last-child td {{ border-bottom: none; }}
  .game-row.win  {{ background: var(--win-bg); }}
  .game-row.loss {{ background: var(--loss-bg); }}
  .game-row.draw {{ background: var(--draw-bg); }}
  .bd-score  {{ font-size: 16px; font-weight: 900; color: var(--primary); }}
  .opp-score {{ font-size: 16px; color: var(--muted); }}
  .result-badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
                   font-size: 12px; font-weight: 700; }}
  .result-badge.win  {{ background: var(--win-cl);  color: #fff; }}
  .result-badge.loss {{ background: var(--loss-cl); color: #fff; }}
  .result-badge.draw {{ background: var(--draw-cl); color: #fff; }}
  .inn-cell {{ font-size: 11px; color: var(--muted); width: 26px; }}
  .tbl-wrap {{ overflow-x: auto; background: var(--card-bg); border: 1px solid var(--border);
               border-radius: 10px; overflow: hidden; }}
  table.stats-tbl {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  table.stats-tbl th {{ background: var(--primary); color: #fff; padding: 7px 9px;
                         text-align: center; white-space: nowrap; }}
  table.stats-tbl td {{ padding: 6px 9px; border-bottom: 1px solid #ede8dc; text-align: center; }}
  table.stats-tbl tr:last-child td {{ border-bottom: none; }}
  .name-col {{ text-align: left !important; font-weight: 700; }}
  .hit-col  {{ font-weight: 700; }}
  .rate-col {{ font-weight: 700; color: var(--primary); }}
  .ops-col  {{ color: var(--accent) !important; }}
  .ip-col   {{ white-space: nowrap; }}
  .unqual td {{ color: var(--muted); }}
  .unqual .name-col {{ color: var(--text); font-weight: 600; }}
  .section-divider td {{ font-size: 11px; font-weight: 700; color: var(--muted);
                          background: #f0ece0 !important; padding: 5px 9px !important; }}
  .qual-badge {{ font-size: 9px; font-weight: 700; background: var(--gold); color: #fff;
                 border-radius: 3px; padding: 1px 4px; margin-left: 4px; vertical-align: middle; }}
  .rank-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(160px,1fr)); gap: 12px; margin-bottom: 32px; }}
  .rank-card {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }}
  .rank-hdr  {{ background: var(--primary); color: #fff; font-size: 12px; font-weight: 700; padding: 8px 12px; }}
  .rank-row  {{ display: flex; align-items: center; padding: 7px 10px; gap: 8px;
                border-bottom: 1px solid var(--border); font-size: 13px; }}
  .rank-row:last-child {{ border-bottom: none; }}
  .rank-row.rank-1 {{ background: #ddd; color: var(--muted); }}
  .rank-num  {{ width: 18px; font-size: 11px; font-weight: 700; color: var(--muted); text-align: center; }}
  .rank-1 .rank-num {{ background: var(--gold); color: #fff; border-radius: 3px; }}
  .rank-name {{ flex: 1; font-weight: 600; }}
  .rank-val  {{ font-weight: 900; color: var(--primary); font-size: 13px; }}
  .rank-1 {{ border-bottom: 1px dashed var(--primary); }}
  .rank-1 .rank-name, .rank-1 .rank-val {{ color: var(--primary); background: rgba(26,58,92,.06); }}
  .game-link, .player-link {{ color: inherit; text-decoration: none; border-bottom: 1px dotted var(--primary); cursor: pointer; }}
  .modal-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,.5);
                    z-index: 1000; align-items: center; justify-content: center; padding: 20px; }}
  .modal-overlay.open {{ display: flex; }}
  .modal {{ background: var(--card-bg); border-radius: 14px; max-width: 900px; width: 100%;
            max-height: 85vh; overflow-y: auto; }}
  .modal-header {{ background: var(--primary); color: #fff; padding: 16px 20px;
                   border-radius: 14px 14px 0 0; display: flex; justify-content: space-between; align-items: center; }}
  .modal-title  {{ font-size: 16px; font-weight: 900; }}
  .modal-close  {{ background: none; border: none; color: #fff; font-size: 22px; cursor: pointer; opacity: .7; }}
  .modal-close:hover {{ opacity: 1; }}
  .modal-body   {{ padding: 16px 20px; }}
  .modal-section {{ margin-bottom: 20px; }}
  .modal-sec-title {{ font-size: 13px; font-weight: 900; color: var(--primary);
                      margin-bottom: 8px; padding-bottom: 6px; border-bottom: 2px solid var(--primary); }}
  .modal-tbl-wrap {{ overflow-x: auto; }}
  table.modal-tbl {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  table.modal-tbl th {{ background: var(--primary); color: #fff; padding: 6px 8px; text-align: center; }}
  table.modal-tbl td {{ padding: 5px 8px; border-bottom: 1px solid #ede8dc; text-align: center; }}
  table.modal-tbl .name-col {{ text-align: left !important; font-weight: 700; }}
  .no-data {{ color: var(--muted); padding: 12px; text-align: center; font-style: italic; }}
  .year-nav {{ background: #0d2238; padding: 6px 24px; display: flex; align-items: center;
               gap: 6px; flex-wrap: wrap; border-bottom: 1px solid rgba(255,255,255,.08); }}
  .year-nav-label {{ font-size: 11px; color: #6a8aaa; letter-spacing: .1em; margin-right: 4px; }}
  .year-nav a {{ font-size: 12px; color: #7a9ab8; text-decoration: none; padding: 3px 9px;
                 border-radius: 3px; border: 1px solid rgba(255,255,255,.1); transition: all .15s; }}
  .year-nav a:hover {{ background: rgba(255,255,255,.1); color: #fff; }}
  .year-nav a.current {{ background: var(--gold); color: #1a2030; font-weight: 700;
                          border-color: var(--gold); }}
  @media (max-width:600px) {{
    .hero {{ padding: 24px 16px 20px; }} .main {{ padding: 20px 16px 40px; }}
  }}
</style>
</head>
<body>

<nav class="nav">
  <a href="index.html">🏠 トップ</a>
  <a href="times.html">📰 タイムズ</a>
  <a href="players.html">⚾ 選手成績</a>
  <a href="records.html">🏆 通算記録</a>
  <a href="season.html" class="active">📋 今シーズン</a>
  <a href="titles.html">🎖 タイトル</a>
</nav>
<div class="year-nav">
  <span class="year-nav-label">年度：</span>
  {year_links}
</div>

<div class="hero">
  <div class="hero-year">{year} SEASON</div>
  <h1 class="hero-title">バミューダ<span>ドルフィンズ</span><br>{year}年シーズン成績</h1>
  <div class="rec-row">
    <div class="rec-item"><div class="rec-num rec-win">{wins}</div><div class="rec-lbl">勝</div></div>
    <div class="rec-item"><div class="rec-num rec-loss">{losses}</div><div class="rec-lbl">敗</div></div>
    <div class="rec-item"><div class="rec-num rec-draw">{draws}</div><div class="rec-lbl">分</div></div>
    <div class="rec-item"><div class="rec-num">{total}</div><div class="rec-lbl">試合</div></div>
  </div>
</div>

<div class="main">

  <!-- チーム成績 -->
  <div class="sec-hdr">📊 チーム成績</div>
  <div class="team-card">
    <div class="tc-head">🏏 チーム打撃</div>
    <div class="tc-stats">
      <div class="tc-stat"><div class="tc-val">{tb_r}</div><div class="tc-lbl">得点</div></div>
      <div class="tc-stat"><div class="tc-val">{tb_h}</div><div class="tc-lbl">安打</div></div>
      <div class="tc-stat"><div class="tc-val">{tb_rbi}</div><div class="tc-lbl">打点</div></div>
      <div class="tc-stat"><div class="tc-val">{tb_sb}</div><div class="tc-lbl">盗塁</div></div>
      <div class="tc-stat"><div class="tc-val">{tb_bb}</div><div class="tc-lbl">四死球</div></div>
      <div class="tc-stat"><div class="tc-val">{tb_so}</div><div class="tc-lbl">三振</div></div>
      <div class="tc-stat"><div class="tc-rate">{avg}</div><div class="tc-lbl">打率</div></div>
      <div class="tc-stat"><div class="tc-rate">{obp}</div><div class="tc-lbl">出塁率</div></div>
      <div class="tc-stat"><div class="tc-rate">{slg}</div><div class="tc-lbl">長打率</div></div>
    </div>
  </div>
  <div class="team-card">
    <div class="tc-head">⚾ 投手成績</div>
    <div class="tc-stats">
      <div class="tc-stat"><div class="tc-val">{tp_w}</div><div class="tc-lbl">勝利</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_l}</div><div class="tc-lbl">敗北</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_s}</div><div class="tc-lbl">セーブ</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_ip}</div><div class="tc-lbl">投球回</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_h}</div><div class="tc-lbl">被安打</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_bb}</div><div class="tc-lbl">四死球</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_so}</div><div class="tc-lbl">奪三振</div></div>
      <div class="tc-stat"><div class="tc-val">{tp_r}</div><div class="tc-lbl">失点</div></div>
      <div class="tc-stat"><div class="tc-rate">{era}</div><div class="tc-lbl">防御率</div></div>
    </div>
  </div>

  <!-- 試合結果 -->
  <div class="sec-hdr">📅 試合結果</div>
  <div class="games-wrap">
    <table class="games-tbl">
      <thead>
        <tr>
          <th>日付</th><th>対戦相手</th>
          <th colspan="3">スコア</th><th>結果</th>
          <th>1</th><th>2</th><th>3</th><th>4</th><th>5</th>
          <th>6</th><th>7</th><th>8</th><th>9</th>
        </tr>
      </thead>
      <tbody>{game_rows}</tbody>
    </table>
  </div>

  <!-- 打撃成績 -->
  <div class="sec-hdr">🏏 打撃成績 <small style="font-size:12px;font-weight:400;color:var(--muted)">選手名をクリックで試合別詳細</small></div>
  <div class="tbl-wrap">
    <table class="stats-tbl">
      <thead>
        <tr>
          <th>選手</th><th>試</th><th>打席</th><th>打数</th><th>得</th>
          <th>安打</th><th>打点</th><th>盗塁</th>
          <th>二塁</th><th>三塁</th><th>本塁</th>
          <th>四死</th><th>三振</th>
          <th>打率</th><th>出塁率</th><th>長打率</th><th>OPS</th>
        </tr>
      </thead>
      <tbody>{bat_rows}</tbody>
    </table>
  </div>

  <!-- 部門別ランキング -->
  <div class="sec-hdr">🏅 部門別ランキング</div>
  {rank_sections}

  <!-- 投手成績 -->
  <div class="sec-hdr">⚾ 投手成績 <small style="font-size:12px;font-weight:400;color:var(--muted)">選手名をクリックで試合別詳細</small></div>
  <div class="tbl-wrap">
    <table class="stats-tbl">
      <thead>
        <tr>
          <th>投手</th><th>試</th><th>勝</th><th>敗</th><th>S</th>
          <th>投球回</th><th>打者</th><th>被安</th><th>四死</th><th>奪三</th>
          <th>失点</th><th>自責</th><th>防御率</th>
        </tr>
      </thead>
      <tbody>{pit_rows}</tbody>
    </table>
  </div>

</div>

<!-- Modal -->
<div class="modal-overlay" id="modal" onclick="closeIfOverlay(event)">
  <div class="modal">
    <div class="modal-header">
      <div class="modal-title" id="modal-title">詳細</div>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="modal-body" id="modal-body"></div>
  </div>
</div>

<script>
const oppDate   = {opp_date_json};
const gameBat   = {game_bat_json};
const gamePit   = {game_pit_json};
const playerBat = {player_bat_json};
const playerPit = {player_pit_json};

function openModal(title, html) {{
  document.getElementById('modal-title').textContent = title;
  document.getElementById('modal-body').innerHTML = html;
  document.getElementById('modal').classList.add('open');
}}
function closeModal() {{ document.getElementById('modal').classList.remove('open'); }}
function closeIfOverlay(e) {{ if (e.target === document.getElementById('modal')) closeModal(); }}
document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closeModal(); }});

function td(v) {{ return `<td>${{v ?? ''}}</td>`; }}
function fmtRate(v) {{
  if (isNaN(v)) return '—';
  const s = v.toFixed(3);
  return s.startsWith('0.') ? s.slice(1) : s;
}}
function ipToNum(ip) {{
  if (!ip) return 0;
  if (ip.includes('1/3')) return parseInt(ip) + 1/3;
  if (ip.includes('2/3')) return parseInt(ip) + 2/3;
  return parseInt(ip)||0;
}}

function showPlayer(e, name) {{
  e.preventDefault();
  const bRecs = playerBat[name] || [];
  const pRecs = playerPit[name] || [];
  let html = '';
  if (bRecs.length) {{
    let cAb=0, cH=0;
    html += `<div class="modal-section"><div class="modal-sec-title">🏏 打撃成績（試合別）</div>
      <div class="modal-tbl-wrap"><table class="modal-tbl">
      <thead><tr><th>日付</th><th>対戦相手</th><th>打席</th><th>打数</th><th>得</th>
      <th>安打</th><th>打点</th><th>盗塁</th><th>二塁</th><th>三塁</th><th>本塁</th>
      <th>四死</th><th>三振</th><th>打率</th></tr></thead><tbody>`;
    for (const r of bRecs) {{
      cAb += parseInt(r.ab)||0; cH += parseInt(r.h)||0;
      const avg = cAb>0 ? fmtRate(cH/cAb) : '—';
      html += `<tr><td>${{r.date}}</td><td class="name-col">${{r.opp}}</td>
        ${{td(r.pa)}}${{td(r.ab)}}${{td(r.r)}}${{td(r.h)}}${{td(r.rbi)}}${{td(r.sb)}}
        ${{td(r.h2)}}${{td(r.h3)}}${{td(r.hr)}}${{td(r.bb)}}${{td(r.so)}}
        <td class="rate-col">${{avg}}</td></tr>`;
    }}
    html += '</tbody></table></div></div>';
  }}
  if (pRecs.length) {{
    let cIp=0, cEr=0;
    html += `<div class="modal-section"><div class="modal-sec-title">⚾ 投手成績（試合別）</div>
      <div class="modal-tbl-wrap"><table class="modal-tbl">
      <thead><tr><th>日付</th><th>対戦相手</th><th>結果</th><th>投球回</th>
      <th>打者</th><th>被安</th><th>四死</th><th>奪三</th><th>失点</th><th>自責</th>
      <th>防御率</th></tr></thead><tbody>`;
    for (const r of pRecs) {{
      cIp += ipToNum(r.ip); cEr += parseInt(r.er)||0;
      const era = cIp>0 ? fmtRate(cEr*9/cIp) : '—';
      html += `<tr><td>${{r.date}}</td><td class="name-col">${{r.opp}}</td>
        ${{td(r.wl)}}${{td(r.ip)}}${{td(r.bf)}}${{td(r.h)}}${{td(r.bb)}}
        ${{td(r.so)}}${{td(r.r)}}${{td(r.er)}}<td class="rate-col">${{era}}</td></tr>`;
    }}
    html += '</tbody></table></div></div>';
  }}
  if (!html) html = '<p class="no-data">記録がありません</p>';
  openModal(`${{name}} 個人成績`, html);
}}

function showGame(e, opp, date) {{
  e.preventDefault();
  const key = opp + '|' + date;
  const bats = gameBat[key] || [];
  const pits = gamePit[key] || [];
  let html = '';
  if (bats.length) {{
    let cums = {{}};
    html += `<div class="modal-section"><div class="modal-sec-title">🏏 打撃成績</div>
      <div class="modal-tbl-wrap"><table class="modal-tbl">
      <thead><tr><th>選手</th><th>打席</th><th>打数</th><th>得</th><th>安打</th>
      <th>打点</th><th>盗塁</th><th>二塁</th><th>三塁</th><th>本塁</th>
      <th>四死</th><th>三振</th><th>今季打率</th></tr></thead><tbody>`;
    for (const r of bats) {{
      const allG = playerBat[r.name] || [];
      let cAb=0, cH=0;
      for (const g of allG) {{ cAb+=parseInt(g.ab)||0; cH+=parseInt(g.h)||0; if(g.opp===opp&&g.date===date) break; }}
      const avg = cAb>0 ? fmtRate(cH/cAb) : '—';
      html += `<tr><td class="name-col">${{r.name}}</td>
        ${{td(r.pa)}}${{td(r.ab)}}${{td(r.r)}}${{td(r.h)}}${{td(r.rbi)}}${{td(r.sb)}}
        ${{td(r.h2)}}${{td(r.h3)}}${{td(r.hr)}}${{td(r.bb)}}${{td(r.so)}}
        <td class="rate-col">${{avg}}</td></tr>`;
    }}
    html += '</tbody></table></div></div>';
  }}
  if (pits.length) {{
    html += `<div class="modal-section"><div class="modal-sec-title">⚾ 投手成績</div>
      <div class="modal-tbl-wrap"><table class="modal-tbl">
      <thead><tr><th>投手</th><th>結果</th><th>投球回</th><th>打者</th>
      <th>被安</th><th>四死</th><th>奪三</th><th>失点</th><th>自責</th>
      <th>今季防御率</th></tr></thead><tbody>`;
    for (const r of pits) {{
      const allP = playerPit[r.name] || [];
      let cIp=0, cEr=0;
      for (const g of allP) {{ cIp+=ipToNum(g.ip); cEr+=parseInt(g.er)||0; if(g.opp===opp&&g.date===date) break; }}
      const era = cIp>0 ? fmtRate(cEr*9/cIp) : '—';
      html += `<tr><td class="name-col">${{r.name}}</td>${{td(r.wl)}}
        ${{td(r.ip)}}${{td(r.bf)}}${{td(r.h)}}${{td(r.bb)}}
        ${{td(r.so)}}${{td(r.r)}}${{td(r.er)}}<td class="rate-col">${{era}}</td></tr>`;
    }}
    html += '</tbody></table></div></div>';
  }}
  if (!html) html = '<p class="no-data">記録がありません</p>';
  openModal(`${{date}} vs ${{opp}} 個人成績`, html);
}}
</script>
</body>
</html>
"""

# ─── メイン ────────────────────────────────────────────────────────────────

KNOWN_YEARS = [2026, 2025]  # ← 新しい年度が増えたらここに追加

def make_year_links(current_year):
    links = []
    for y in sorted(KNOWN_YEARS, reverse=True):
        cls = ' class="current"' if str(y) == str(current_year) else ''
        fname = "season.html" if y == max(KNOWN_YEARS) else f"season_{y}.html"
        links.append(f'<a href="{fname}"{cls}>{y}年</a>')
    return "\n  ".join(links)

def main():
    if len(sys.argv) < 2:
        print("使い方: python generate_season.py <Excelファイルパス> [年度]")
        print("例:     python generate_season.py ../Bermudas_works/Bmds_shukei_2026.xlsx 2026")
        print("例:     python generate_season.py ../Bermudas_works/Bmds_shukei_2025.xlsx 2025")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    year      = sys.argv[2] if len(sys.argv) > 2 else str(max(KNOWN_YEARS))
    # 最新年度はseason.html、過去年度はseason_YYYY.html
    out_name  = "season.html" if int(year) == max(KNOWN_YEARS) else f"season_{year}.html"

    print(f"📂 読み込み中: {xlsx_path}")
    games, game_bat, player_bat, game_pit, player_pit = load_excel(xlsx_path)
    print(f"✅ 試合数: {len(games)}, 打撃レコード: {sum(len(v) for v in game_bat.values())}, 投手レコード: {sum(len(v) for v in game_pit.values())}")

    wins, losses, draws, total, tb, avg, obp, slg, tp, era = compute_team_stats(games, game_bat, game_pit)
    bat_stats = compute_player_bat_stats(player_bat, total)
    pit_stats = compute_player_pit_stats(player_pit, total)

    opp_date_json   = json.dumps({g["opp"]: g["date"] for g in games}, ensure_ascii=False)
    game_bat_json   = json.dumps(game_bat,   ensure_ascii=False)
    game_pit_json   = json.dumps(game_pit,   ensure_ascii=False)
    player_bat_json = json.dumps(player_bat, ensure_ascii=False)
    player_pit_json = json.dumps(player_pit, ensure_ascii=False)

    html = HTML_TEMPLATE.format(
        year=year,
        year_links=make_year_links(year),
        wins=wins, losses=losses, draws=draws, total=total,
        tb_r=tb["r"], tb_h=tb["h"], tb_rbi=tb["rbi"], tb_sb=tb["sb"],
        tb_bb=tb["bb"], tb_so=tb["so"],
        avg=avg, obp=obp, slg=slg,
        tp_w=tp["w"], tp_l=tp["l"], tp_s=tp["s"],
        tp_ip=num_to_ip_str(tp["ip"]), tp_h=tp["h"],
        tp_bb=tp["bb"], tp_so=tp["so"], tp_r=tp["r"], era=era,
        game_rows=game_rows_html(games),
        bat_rows=bat_rows_html(bat_stats, total),
        pit_rows=pit_rows_html(pit_stats, total),
        rank_sections=ranking_sections_html(bat_stats, pit_stats),
        opp_date_json=opp_date_json,
        game_bat_json=game_bat_json,
        game_pit_json=game_pit_json,
        player_bat_json=player_bat_json,
        player_pit_json=player_pit_json,
    )

    out_path = Path(__file__).parent / out_name
    out_path.write_text(html, encoding="utf-8")
    print(f"✅ 生成完了: {out_path}")
    if out_name != "season.html":
        print(f"ℹ️  season.html は最新年度({max(KNOWN_YEARS)}年)が自動的に使用されます")

if __name__ == "__main__":
    main()
